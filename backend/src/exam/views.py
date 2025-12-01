from django.shortcuts import get_object_or_404
from django.http import HttpResponse
from django.utils import timezone
from django.db.models import Q


from rest_framework import viewsets, status, permissions as drf_permissions,filters
from rest_framework.views import APIView
from rest_framework.decorators import action
from rest_framework.response import Response

import openpyxl
from openpyxl.utils import get_column_letter
from .permissions import IsAdmin
# ============================
# IMPORT MODELS
# ============================
from .models import (
    Course,
    CourseParticipant,
    CourseSyllabus,
    CourseMaterial,
    Exam,
    Question,
    Choice,
    UserExam,
    UserAnswer,
    CourseTask,
    CourseTaskSubmission,
    CourseTaskSubmissionFile,
    CourseRequirementAnswer,
    CourseRequirementSubmission,
    CourseRequirementTemplate,
    CourseAssessment,
    CourseAssessmentCriteria,
    UserAnswerFile
)

# ============================
# IMPORT SERIALIZERS
# ============================
from .serializers import (
    CourseSerializer,
    CourseParticipantSerializer,
    CourseSyllabusSerializer,
    CourseSyllabusCreateUpdateSerializer,
    CourseJoinSerializer,
    AssignRoleSerializer,

    CourseMaterialSerializer,
    CourseMaterialCreateUpdateSerializer,

    ExamAdminSerializer,
    ExamPublicSerializer,
    ExamResultSerializer,

    QuestionCreateUpdateSerializer,
    QuestionPublicSerializer,

    SubmitAnswerSerializer,
    QuestionAdminSerializer,

    CourseTaskSerializer,
    CourseTaskSubmissionSerializer,
    CourseTaskSubmissionFileSerializer,
    CoursePublicSerializer,

    CourseRequirementTemplateSerializer,
    CourseRequirementSubmissionSerializer,
    CourseRequirementAnswerSerializer,
    CourseRequirementSubmission,
    CourseRequirementTemplate,  

    CourseAssessmentCriteriaSerializer,
    CourseAssessmentCreateSerializer,
    CourseAssessmentSerializer
)

# ============================
# IMPORT PERMISSIONS
# ============================
from .permissions import (
    IsAdmin,
    IsTrainer,
    IsAssessor,
    IsTrainerOrAssessor,
    IsCourseParticipant,
    IsExamParticipant,
    IsExamInstructorOrAssessor,
    IsTaskSubmissionOwner,
    IsTaskGrader,
    IsTrainerOrAdmin,
    IsExamCreator,
    user_role_in_course,
)


# ================================================================
# ⚡  COURSE VIEWSET
# ================================================================
class CourseViewSet(viewsets.ModelViewSet):
    queryset = Course.objects.all().order_by("-created_at")
    serializer_class = CourseSerializer
    permission_classes = [drf_permissions.IsAuthenticated]
    filter_backends = [filters.SearchFilter]
    search_fields = ["title", "description"]

    def get_permissions(self):
        if self.action in ["create", "update", "partial_update", "destroy"]:
            return [IsTrainerOrAdmin()]
        return [drf_permissions.IsAuthenticated()]

    # =====================================================================
    # JOIN COURSE
    # =====================================================================
    @action(detail=True, methods=["post"], url_path="join")
    def join(self, request, pk=None):
        course = self.get_object()

        # jika course punya requirement → harus isi persyaratan
        if course.requirements.exists():
            return Response({
                "detail": "Course ini memerlukan persyaratan. Silakan isi form persyaratan.",
                "requires_approval": True
            }, status=400)

        # Jika tidak ada requirement → join dengan token
        serializer = CourseJoinSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        if serializer.validated_data["token"] != course.token:
            return Response({"detail": "Token salah."}, status=400)

        cp, created = CourseParticipant.objects.get_or_create(
            user=request.user,
            course=course
        )

        if not created:
            return Response({"detail": "Anda sudah terdaftar."}, status=400)

        return Response({"detail": "Berhasil join.", "participant_id": cp.id})

    # =====================================================================
    # LIST PARTICIPANTS
    # =====================================================================
    @action(detail=True, methods=["get"], url_path="participants")
    def participants(self, request, pk=None):
        course = self.get_object()
        qs = CourseParticipant.objects.filter(course=course)
        return Response(CourseParticipantSerializer(qs, many=True).data)

    # =====================================================================
    # ASSIGN ROLE
    # =====================================================================
    @action(detail=True, methods=["post"], url_path="assign-role")
    def assign_role(self, request, pk=None):
        course = self.get_object()

        if not (request.user.is_staff or user_role_in_course(request.user, course.id, ["trainer"])):
            return Response({"detail": "Tidak diizinkan."}, status=403)

        ser = AssignRoleSerializer(data=request.data)
        ser.is_valid(raise_exception=True)

        user_id = ser.validated_data["user_id"]
        role = ser.validated_data["role"]

        cp, _ = CourseParticipant.objects.get_or_create(user_id=user_id, course=course)
        cp.role = role
        cp.save()

        return Response({"detail": "Role diperbarui."})

    # =====================================================================
    # REQUIREMENTS — LIST TEMPLATE + USER SUBMISSION
    # =====================================================================
    @action(detail=True, methods=["get"], url_path="requirements")
    def list_requirements(self, request, pk=None):
        course = self.get_object()

        templates = CourseRequirementTemplate.objects.filter(course=course).order_by("order")
        temp_ser = CourseRequirementTemplateSerializer(templates, many=True).data

        # get latest submission per user
        latest = CourseRequirementSubmission.objects.filter(
            course=course,
            user=request.user
        ).order_by("-submitted_at").first()

        submission_data = None

        if latest:
            answers = CourseRequirementAnswer.objects.filter(submission=latest)
            submission_data = {
                "id": latest.id,
                "status": latest.status,
                "submitted_at": latest.submitted_at,
                "reviewed_at": latest.reviewed_at,
                "reviewer": latest.reviewer.id if latest.reviewer else None,
                "note": latest.note,
                "answers": CourseRequirementAnswerSerializer(answers, many=True).data
            }

        return Response({
            "templates": temp_ser,
            "user_submission": submission_data
        })

    # =====================================================================
    # SUBMIT REQUIREMENTS  (file upload supported)
    # =====================================================================
    @action(detail=True, methods=["post"], url_path="requirements/submit")
    def submit_requirements(self, request, pk=None):
        course = self.get_object()

        if not course.requirements.exists():
            return Response({"detail": "Course ini tidak memiliki persyaratan."}, status=400)

        import json

        # If multipart (files included)
        if request.content_type.startswith("multipart/"):
            answers_json = request.POST.get("answers")
            if not answers_json:
                return Response({"detail": "Data 'answers' tidak dikirim."}, status=400)
            try:
                answers_list = json.loads(answers_json)
            except:
                return Response({"detail": "Format JSON jawaban invalid."}, status=400)

            # attach file(s)
            for ans in answers_list:
                req_id = str(ans.get("requirement"))
                f_key = f"file_{req_id}"
                if f_key in request.FILES:
                    ans["value_file"] = request.FILES[f_key]
        else:
            answers_list = request.data.get("answers") or []

        if not answers_list:
            return Response({"detail": "Tidak ada jawaban dikirim."}, status=400)

        # create submission
        submission = CourseRequirementSubmission.objects.create(
            course=course,
            user=request.user,
            status="pending"
        )

        created_ids = []
        for ans in answers_list:
            req = CourseRequirementTemplate.objects.filter(
                id=ans.get("requirement"),
                course=course
            ).first()
            if not req:
                continue

            obj = CourseRequirementAnswer.objects.create(
                submission=submission,
                requirement=req,
                value_text=ans.get("value_text"),
                value_number=ans.get("value_number"),
                value_file=ans.get("value_file", None)
            )
            created_ids.append(obj.id)

        return Response({
            "detail": "Persyaratan berhasil diajukan.",
            "submission_id": submission.id,
            "answers_created": created_ids
        }, status=201)

    # =====================================================================
    # REQUIREMENTS — ADMIN LIST SUBMISSIONS
    # =====================================================================
    @action(detail=True, methods=["get"], url_path="submissions")
    def submissions(self, request, pk=None):
        course = self.get_object()

        if not request.user.is_staff:
            return Response({"detail": "Tidak diizinkan."}, status=403)

        subs = CourseRequirementSubmission.objects.filter(course=course).order_by("-submitted_at")

        data = []
        for s in subs:
            answers = CourseRequirementAnswer.objects.filter(submission=s)
            data.append({
                "id": s.id,
                "user": s.user.username,
                "status": s.status,
                "submitted_at": s.submitted_at,
                "reviewed_at": s.reviewed_at,
                "reviewer": s.reviewer.username if s.reviewer else None,
                "answers": CourseRequirementAnswerSerializer(answers, many=True).data
            })

        return Response(data)

    # =====================================================================
    # REQUIREMENTS — APPROVE
    # =====================================================================
    @action(detail=True, methods=["patch"], url_path="submission/(?P<sid>[^/.]+)/approve")
    def approve_submission(self, request, pk=None, sid=None):
        course = self.get_object()

        if not request.user.is_staff:
            return Response({"detail": "Tidak diizinkan."}, status=403)

        submission = get_object_or_404(CourseRequirementSubmission, id=sid, course=course)

        submission.status = "approved"
        submission.reviewed_at = timezone.now()
        submission.reviewer = request.user
        submission.save()

        # auto-add participant
        CourseParticipant.objects.get_or_create(
            course=course,
            user=submission.user
        )

        return Response({"detail": "Submission disetujui dan user menjadi peserta."})

    # =====================================================================
    # REQUIREMENTS — REJECT
    # =====================================================================
    @action(detail=True, methods=["patch"], url_path="submission/(?P<sid>[^/.]+)/reject")
    def reject_submission(self, request, pk=None, sid=None):
        course = self.get_object()

        if not request.user.is_staff:
            return Response({"detail": "Tidak diizinkan."}, status=403)

        submission = get_object_or_404(CourseRequirementSubmission, id=sid, course=course)

        submission.status = "rejected"
        submission.note = request.data.get("note", "")
        submission.reviewed_at = timezone.now()
        submission.reviewer = request.user
        submission.save()

        return Response({"detail": "Submission ditolak."})

    # =====================================================================
    # REQUIREMENTS — DOWNLOAD FILE
    # =====================================================================
    @action(
        detail=True,
        methods=["get"],
        url_path="submission/(?P<sid>[^/.]+)/download/(?P<answer_id>[^/.]+)",
        permission_classes=[IsAdmin]
    )
    def download_requirement_file(self, request, pk=None, sid=None, answer_id=None):

        submission = get_object_or_404(
            CourseRequirementSubmission, id=sid, course_id=pk
        )
        answer = get_object_or_404(
            CourseRequirementAnswer, id=answer_id, submission=submission
        )

        if not answer.value_file:
            return Response({"detail": "Tidak ada file."}, status=404)

        file_name = answer.value_file.name.split("/")[-1]

        response = HttpResponse(
            answer.value_file.open("rb").read(),
            content_type="application/octet-stream"
        )
        response["Content-Disposition"] = f'attachment; filename="{file_name}"'
        return response

    # =====================================================================
    # SYLLABUS CRUD
    # =====================================================================
    def get_serializer_class(self):
        if self.action in ["list", "retrieve"]:
            return CoursePublicSerializer

        if self.action in ["syllabus_create", "syllabus_update"]:
            return CourseSyllabusCreateUpdateSerializer

        return CourseSerializer



    @action(detail=True, methods=["get"], url_path="syllabus")
    def list_syllabus(self, request, pk=None):
        course = self.get_object()
        syllabus = course.syllabus.all().order_by("id")
        serializer = CourseSyllabusSerializer(syllabus, many=True)
        return Response(serializer.data)

    @action(detail=True, methods=["post"], url_path="syllabus/create")
    def syllabus_create(self, request, pk=None):
        course = self.get_object()
        if not (request.user.is_staff or user_role_in_course(request.user, course.id, ["trainer"])):
            return Response({"detail": "Tidak diizinkan."}, status=403)

        serializer = CourseSyllabusCreateUpdateSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        syllabus = serializer.save(course=course)

        return Response(CourseSyllabusSerializer(syllabus).data, status=201)

    @action(detail=True, methods=["patch"], url_path="syllabus/(?P<sid>[^/.]+)/update")
    def syllabus_update(self, request, pk=None, sid=None):
        course = self.get_object()
        if not (request.user.is_staff or user_role_in_course(request.user, course.id, ["trainer"])):
            return Response({"detail": "Tidak diizinkan."}, status=403)

        syllabus = get_object_or_404(CourseSyllabus, id=sid, course=course)
        ser = CourseSyllabusCreateUpdateSerializer(syllabus, data=request.data, partial=True)
        ser.is_valid(raise_exception=True)
        ser.save()

        return Response(CourseSyllabusSerializer(syllabus).data)

    @action(detail=True, methods=["delete"], url_path="syllabus/(?P<sid>[^/.]+)/delete")
    def syllabus_delete(self, request, pk=None, sid=None):
        course = self.get_object()
        if not (request.user.is_staff or user_role_in_course(request.user, course.id, ["trainer"])):
            return Response({"detail": "Tidak diizinkan."}, status=403)

        syllabus = get_object_or_404(CourseSyllabus, id=sid, course=course)
        syllabus.delete()
        return Response({"detail": "Silabus dihapus."})

    # =====================================================================
    # TASK LIST
    # =====================================================================
    @action(detail=True, methods=["get"], url_path="tasks")
    def list_tasks(self, request, pk=None):
        course = self.get_object()
        tasks = course.tasks.all().order_by("-created_at")
        serializer = CourseTaskSerializer(tasks, many=True)
        return Response(serializer.data)

    # =====================================================================
    # MATERIAL CRUD
    # =====================================================================
    @action(detail=True, methods=["get"], url_path="materials")
    def list_materials(self, request, pk=None):
        course = self.get_object()
        materials = course.materials.all().order_by("id")
        serializer = CourseMaterialSerializer(materials, many=True)
        return Response(serializer.data)

    @action(detail=True, methods=["post"], url_path="materials/create")
    def create_material(self, request, pk=None):
        course = self.get_object()

        if not (request.user.is_staff or user_role_in_course(request.user, course.id, ["trainer"])):
            return Response({"detail": "Tidak diizinkan."}, status=403)

        ser = CourseMaterialCreateUpdateSerializer(data=request.data)
        ser.is_valid(raise_exception=True)
        mat = ser.save(course=course)

        return Response(CourseMaterialSerializer(mat).data, status=201)

    @action(detail=True, methods=["patch"], url_path="materials/(?P<mid>[^/.]+)/update")
    def update_material(self, request, pk=None, mid=None):
        course = self.get_object()

        if not (request.user.is_staff or user_role_in_course(request.user, course.id, ["trainer"])):
            return Response({"detail": "Tidak diizinkan."}, status=403)

        mat = get_object_or_404(CourseMaterial, id=mid, course=course)
        ser = CourseMaterialCreateUpdateSerializer(mat, data=request.data, partial=True)
        ser.is_valid(raise_exception=True)
        ser.save()

        return Response(CourseMaterialSerializer(mat).data)

    @action(detail=True, methods=["delete"], url_path="materials/(?P<mid>[^/.]+)/delete")
    def delete_material(self, request, pk=None, mid=None):
        course = self.get_object()

        if not (request.user.is_staff or user_role_in_course(request.user, course.id, ["trainer"])):
            return Response({"detail": "Tidak diizinkan."}, status=403)

        mat = get_object_or_404(CourseMaterial, id=mid, course=course)
        mat.delete()
        return Response({"detail": "Materi dihapus."})

    # =====================================================================
    # LIST EXAMS
    # =====================================================================
    @action(detail=True, methods=["get"], url_path="exams")
    def list_exams(self, request, pk=None):
        course = self.get_object()
        exams = course.exams.all().order_by("-created_at")
        serializer = ExamPublicSerializer(exams, many=True, context={"request": request})
        return Response(serializer.data)

    # =====================================================================
    # ASSESSMENT & EVALUATION
    # =====================================================================
    @action(detail=True, methods=["post"], url_path="assessment/criteria/create", permission_classes=[IsAdmin])
    def create_criteria(self, request, pk=None):
        course = self.get_object()
        data = request.data.copy()
        data["course"] = course.id
        serializer = CourseAssessmentCriteriaSerializer(data=data)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(serializer.data, status=201)

    @action(detail=True, methods=["get"], url_path="assessment/criteria")
    def list_criteria(self, request, pk=None):
        course = self.get_object()
        qs = CourseAssessmentCriteria.objects.filter(course=course).order_by("order")
        return Response(CourseAssessmentCriteriaSerializer(qs, many=True).data)

    @action(detail=True, methods=["patch"], url_path="assessment/criteria/(?P<cid>[^/.]+)/update", permission_classes=[IsAdmin])
    def update_criteria(self, request, pk=None, cid=None):
        course = self.get_object()
        crit = get_object_or_404(CourseAssessmentCriteria, id=cid, course=course)
        serializer = CourseAssessmentCriteriaSerializer(crit, data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(serializer.data)

    @action(detail=True, methods=["delete"], url_path="assessment/criteria/(?P<cid>[^/.]+)/delete", permission_classes=[IsAdmin])
    def delete_criteria(self, request, pk=None, cid=None):
        course = self.get_object()
        crit = get_object_or_404(CourseAssessmentCriteria, id=cid, course=course)
        crit.delete()
        return Response({"detail": "deleted"})

    @action(detail=True, methods=["post"], url_path="assessment/submit", permission_classes=[IsTrainer|IsAdmin|IsAssessor])
    def submit_assessment(self, request, pk=None):
        course = self.get_object()
        data = request.data.copy()
        data["course"] = course.id
        serializer = CourseAssessmentCreateSerializer(data=data)
        serializer.is_valid(raise_exception=True)
        assessment = serializer.save(assessor=request.user)
        return Response(CourseAssessmentSerializer(assessment).data, status=201)

    @action(detail=True, methods=["get"], url_path="assessment/(?P<user_id>[^/.]+)")
    def get_assessment(self, request, pk=None, user_id=None):
        course = self.get_object()
        assessment = CourseAssessment.objects.filter(course=course, user__id=user_id).first()
        if not assessment:
            return Response({"detail": "Not found"}, status=404)
        return Response(CourseAssessmentSerializer(assessment).data)

    @action(detail=True, methods=["get"], url_path="evaluation/(?P<user_id>[^/.]+)")
    def evaluation(self, request, pk=None, user_id=None):
        """
        Mode evaluasi final course.
        """
        course = self.get_object()
        mode = course.evaluation_mode or "none"

        exams = []
        mandatory_failed = False
        for exam in course.exams.all():
            ue = exam.userexams.filter(user__id=user_id).order_by("-attempt_number").first()
            score = ue.score if ue else None
            passed = None
            if score is not None and exam.passing_grade is not None:
                passed = (score >= exam.passing_grade)

            exams.append({
                "exam_id": exam.id,
                "title": exam.title,
                "score": score,
                "passing_grade": exam.passing_grade,
                "passed": passed,
                "mandatory": bool(exam.is_mandatory)
            })

            if exam.is_mandatory and passed is False:
                mandatory_failed = True

        assessment = CourseAssessment.objects.filter(course=course, user__id=user_id).first()
        assessment_data = CourseAssessmentSerializer(assessment).data if assessment else None

        final_status = None

        if mode == "none":
            return Response({"evaluation_enabled": False})

        if mode == "exam_only":
            final_status = "passed" if not mandatory_failed else "not_passed"

        elif mode == "assessment_only":
            final_status = assessment.status if assessment else None

        elif mode == "combined":
            if mandatory_failed:
                final_status = "not_passed"
            else:
                final_status = assessment.status if assessment else None

        elif mode == "manual":
            final_status = assessment.status if assessment else None

        return Response({
            "evaluation_enabled": True,
            "mode": mode,
            "exams": exams,
            "assessment": assessment_data,
            "final_status": final_status
        })




# ================================================================
# ⚡  EXAM VIEWSET (SUPER FIXED)
# ================================================================
class ExamViewSet(viewsets.ModelViewSet):
    queryset = Exam.objects.all().order_by("-created_at")
    serializer_class = ExamAdminSerializer
    permission_classes = [drf_permissions.IsAuthenticated]

    # --------------------------------------------
    # PERMISSIONS
    # --------------------------------------------
    def get_permissions(self):
        if self.action in ["start", "questions", "submit", "finish", "my_result"]:
            return [IsCourseParticipant()]

        if self.action in ["list_results", "user_result", "grade_answer", "analytics", "export"]:
            return [IsExamInstructorOrAssessor()]

        if self.action in ["create", "update", "partial_update", "destroy",
                           "create_question", "update_question", "delete_question"]:
            return [IsExamCreator()]

        return [drf_permissions.IsAuthenticated()]

    # --------------------------------------------
    # SAFE SERIALIZER HANDLING
    # --------------------------------------------
    def get_serializer_class(self):
        user = self.request.user

        # LIST
        if self.action == "list":
            return ExamPublicSerializer

        # CRUD
        if self.action in ["create", "update", "partial_update", "destroy"]:
            return ExamAdminSerializer

        # DETAIL
        try:
            exam = self.get_object()
        except:
            exam = None

        if user.is_staff:
            return ExamAdminSerializer

        if exam and user_role_in_course(user, exam.course_id, ["trainer", "assessor"]):
            return ExamAdminSerializer

        return ExamPublicSerializer

    def get_serializer_context(self):
        ctx = super().get_serializer_context()
        ctx["request"] = self.request
        return ctx

    # ============================================================
    # RESULTS
    # ============================================================
    @action(detail=True, methods=["get"], url_path="results")
    def list_results(self, request, pk=None):
        exam = self.get_object()
        results = UserExam.objects.filter(exam=exam)
        return Response(ExamResultSerializer(results, many=True).data)

    @action(detail=True, methods=["get"], url_path="results/(?P<user_id>[^/.]+)")
    def user_result(self, request, pk=None, user_id=None):
        exam = self.get_object()
        result = get_object_or_404(UserExam, exam=exam, user_id=user_id)
        return Response(ExamResultSerializer(result).data)

    @action(detail=True, methods=["get"], url_path="my-result")
    def my_result(self, request, pk=None):
        exam = self.get_object()
        result = UserExam.objects.filter(
            exam=exam,
            user=request.user
        ).order_by("-attempt_number").first()

        if not result:
            return Response({"detail": "Belum mengikuti exam ini."}, status=404)

        return Response(ExamResultSerializer(result).data)

    # ============================================================
    # GRADE ANSWER (ASSESSOR)
    # ============================================================
    @action(detail=True, methods=["post"], url_path="grade-answer/(?P<answer_id>[^/.]+)")
    def grade_answer(self, request, pk=None, answer_id=None):
        exam = self.get_object()
        ans = get_object_or_404(UserAnswer, id=answer_id, user_exam__exam=exam)

        if not CourseParticipant.objects.filter(
            course=exam.course,
            user=request.user,
            role="assessor"
        ).exists() and not request.user.is_staff:
            return Response({"detail": "Tidak diizinkan."}, status=403)

        ans.score = request.data.get("score")
        ans.graded = True
        ans.save()

        return Response({"detail": "Jawaban dinilai."})

    # ============================================================
    # START EXAM
    # ============================================================
    @action(detail=True, methods=["post"], url_path="start")
    def start(self, request, pk=None):
        exam = self.get_object()

        if not exam.is_open():
            return Response({"detail": "Exam tidak aktif."}, status=400)

        if not CourseParticipant.objects.filter(course=exam.course, user=request.user).exists():
            return Response({"detail": "Anda bukan peserta."}, status=403)

        prev = UserExam.objects.filter(user=request.user, exam=exam).count()

        if exam.attempt_limit and prev >= exam.attempt_limit:
            return Response({"detail": "Limit attempt tercapai."}, status=400)

        ue = UserExam.objects.create(
            user=request.user,
            exam=exam,
            attempt_number=prev + 1,
            start_time=timezone.now(),
            status="in_progress",
        )

        return Response({
            "detail": "Exam dimulai.",
            "user_exam_id": ue.id,
            "attempt_number": ue.attempt_number,
        })

    # ============================================================
    # QUESTION CRUD
    # ============================================================
    @action(detail=True, methods=["post"], url_path="questions/create")
    def create_question(self, request, pk=None):
        exam = self.get_object()
        ser = QuestionCreateUpdateSerializer(data=request.data)
        ser.is_valid(raise_exception=True)
        q = Question.objects.create(exam=exam, **ser.validated_data)
        return Response({"question_id": q.id}, status=201)

    @action(detail=True, methods=["patch"], url_path="questions/(?P<qid>[^/.]+)/update")
    def update_question(self, request, pk=None, qid=None):
        exam = self.get_object()
        q = get_object_or_404(Question, id=qid, exam=exam)
        ser = QuestionCreateUpdateSerializer(q, data=request.data, partial=True)
        ser.is_valid(raise_exception=True)
        ser.save()
        return Response({"detail": "Updated."})

    @action(detail=True, methods=["delete"], url_path="questions/(?P<qid>[^/.]+)/delete")
    def delete_question(self, request, pk=None, qid=None):
        exam = self.get_object()
        q = get_object_or_404(Question, id=qid, exam=exam)
        q.delete()
        return Response({"detail": "Deleted."})

    # ============================================================
    # GET QUESTIONS
    # ============================================================
    @action(detail=True, methods=["get"], url_path="questions")
    def questions(self, request, pk=None):
        exam = self.get_object()

        # permission check
        if not CourseParticipant.objects.filter(course=exam.course, user=request.user).exists():
            return Response({"detail": "Tidak diizinkan."}, status=403)

        # Ambil semua pertanyaan exam
        base_qs = exam.questions.all()

        # if exam.shuffle_questions: ... (keep same logic)
        if exam.shuffle_questions:
            base_qs = base_qs.order_by("?")

        # handle random subset (aplikasi perlu hati-hati bila branching ada)
        # but we will compute branching before cutting subset

        # Jika user mengirimkan user_exam (sudah mulai), kita sertakan branch sesuai jawaban user
        user_exam_id = request.query_params.get("user_exam")
        if user_exam_id:
            try:
                ue = UserExam.objects.get(id=user_exam_id, exam=exam, user=request.user)
            except UserExam.DoesNotExist:
                return Response({"detail": "UserExam tidak ditemukan."}, status=404)

            # Ambil semua choice ids yang sudah dipilih di attempt ini
            selected_choice_ids = Choice.objects.filter(
                question__in=exam.questions.all(),
                id__in=UserAnswer.objects.filter(user_exam=ue).values_list('selected_choices', flat=True)
            ).values_list('id', flat=True)

            # Mulai dengan soal-level atas (parent_question is None)
            # plus setiap question whose parent_choice_id is in selected_choice_ids
            allowed_q = base_qs.filter(Q(parent_question__isnull=True) | Q(parent_choice_id__in=selected_choice_ids))

            # Support multi-level chaining: repeat until no new questions
            # (collect ids iteratively)
            allowed_ids = set(allowed_q.values_list("id", flat=True))
            changed = True
            while changed:
                changed = False
                # find questions whose parent_choice is in selected_choice_ids AND parent_question id in allowed_ids
                new_qs = base_qs.filter(parent_choice_id__in=selected_choice_ids, parent_question_id__in=allowed_ids).exclude(id__in=allowed_ids)
                new_ids = set(new_qs.values_list("id", flat=True))
                if new_ids:
                    allowed_ids |= new_ids
                    changed = True

            questions = base_qs.filter(id__in=allowed_ids).order_by("order")
        else:
            # no user_exam: return top-level questions (soal utama) only
            questions = base_qs.filter(parent_question__isnull=True).order_by("order")

        # apply random_question_count AFTER building allowed set if desired
        if exam.random_question_count:
            questions = questions[: exam.random_question_count]

        serializer = QuestionPublicSerializer(questions, many=True)
        return Response(serializer.data)

    # ============================================================
    # SUBMIT ANSWERS
    # ============================================================
    @action(detail=True, methods=["post"], url_path="submit")
    def submit(self, request, pk=None):
        """
        Accepts either:
        - JSON body (application/json) with {"user_exam": id, "answers": [...]}  <-- autosave
        OR
        - multipart/form-data with:
            - "user_exam": id
            - "answers": JSON string (list of answer objects with question, selected_choices, text_answer)
            - files uploaded under keys: files_<question_id> (one or many)
        For file uploads we will attach files to the corresponding UserAnswer via UserAnswerFile.
        """
        exam = self.get_object()
        # parse user_exam (could be in form or json)
        user_exam_id = request.POST.get("user_exam") or request.data.get("user_exam")
        if not user_exam_id:
            return Response({"detail": "user_exam is required."}, status=400)

        ue = get_object_or_404(UserExam, id=user_exam_id, exam=exam, user=request.user)

        if ue.status != "in_progress":
            return Response({"detail": "Exam sudah selesai."}, status=400)

        # Parse answers:
        answers_raw = None

        # If content type is multipart/form-data, DRF might put parsed data in request.data,
        # but files are in request.FILES. We support either JSON body or form with 'answers' JSON string.
        if isinstance(request.data, dict) and "answers" in request.data and not isinstance(request.data["answers"], (list, tuple)):
            # answers possibly a JSON string in multipart form
            try:
                import json
                answers_raw = json.loads(request.data["answers"])
            except Exception:
                return Response({"detail": "Field 'answers' must be valid JSON."}, status=400)
        else:
            # DRF already parsed JSON body into request.data as dict
            answers_raw = request.data.get("answers", None)

        if answers_raw is None:
            return Response({"detail": "No answers provided."}, status=400)

        # At this point answers_raw should be a list of answer dicts.
        for ans in answers_raw:
            qid = ans.get("question")
            if not qid:
                return Response({"detail": "Each answer must include 'question' id."}, status=400)

            q = get_object_or_404(Question, id=qid, exam=exam)

            ua, created = UserAnswer.objects.get_or_create(
                user_exam=ue,
                question=q
            )

            # handle selected_choices (list)
            if "selected_choices" in ans:
                ids = ans.get("selected_choices") or []
                # ensure ints
                try:
                    ids = [int(x) for x in ids]
                except Exception:
                    ids = []
                choices = Choice.objects.filter(id__in=ids, question=q)
                ua.selected_choices.set(choices)

            # handle text_answer
            if "text_answer" in ans:
                ua.text_answer = ans.get("text_answer") or ""

            ua.save()

            # handle file uploads: files expected under request.FILES with key files_<question_id>
            file_field = f"files_{qid}"
            if file_field in request.FILES:
                files = request.FILES.getlist(file_field)
                # create UserAnswerFile entries
                for f in files:
                    UserAnswerFile.objects.create(answer=ua, file=f)

        return Response({"detail": "Jawaban disimpan."})


    # ============================================================
    # FINISH EXAM
    # ============================================================
    @action(detail=True, methods=["post"], url_path="finish")
    def finish(self, request, pk=None):
        exam = self.get_object()

        ue = get_object_or_404(
            UserExam,
            id=request.data.get("user_exam"),
            exam=exam,
            user=request.user
        )

        if ue.status == "completed":
            return Response({"detail": "Sudah selesai."})

        raw = 0
        total = 0

        for ans in ue.answers.select_related("question"):
            q = ans.question

            if q.question_type in ["MCQ", "CHECK", "DROPDOWN", "TRUEFALSE"]:
                score_sum = sum(c.score for c in ans.selected_choices.all())
                ans.score = min(score_sum, q.points)
                ans.graded = True
                ans.save()
                raw += ans.score

            total += q.points

        ue.raw_score = raw
        ue.score = (raw / total) * 100 if total > 0 else 0
        ue.status = "completed"
        ue.end_time = timezone.now()
        ue.finished = True
        ue.save()

        return Response({
            "detail": "Exam selesai.",
            "score": ue.score,
            "raw_score": ue.raw_score
        })
    

    # ============================================================
    # ANALYTICS
    # ============================================================
    @action(detail=True, methods=["get"], url_path="analytics")
    def analytics(self, request, pk=None):
        exam = self.get_object()

        user_exams = UserExam.objects.filter(exam=exam)
        participants = CourseParticipant.objects.filter(course=exam.course, role="participant")
        completed = user_exams.filter(status="completed")

        scores = list(completed.values_list("score", flat=True))
        high = max(scores) if scores else None
        low = min(scores) if scores else None
        avg = sum(scores) / len(scores) if scores else None

        distr = {
            "0-20": sum(1 for s in scores if 0 <= s <= 20),
            "21-40": sum(1 for s in scores if 21 <= s <= 40),
            "41-60": sum(1 for s in scores if 41 <= s <= 60),
            "61-80": sum(1 for s in scores if 61 <= s <= 80),
            "81-100": sum(1 for s in scores if 81 <= s <= 100),
        }

        passed = completed.filter(score__gte=exam.passing_grade).count() if exam.passing_grade else None

        return Response({
            "total_participants": participants.count(),
            "total_attempts": user_exams.count(),
            "completed_count": completed.count(),
            "highest_score": high,
            "lowest_score": low,
            "average_score": avg,
            "score_distribution": distr,
            "passing_grade": exam.passing_grade,
            "passed_count": passed,
        })

    # ============================================================
    # EXPORT EXCEL
    # ============================================================
    @action(detail=True, methods=["get"], url_path="export")
    def export_excel(self, request, pk=None):
        exam = self.get_object()
        user_exams = UserExam.objects.filter(exam=exam).select_related("user")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Exam Results"

        headers = [
            "User ID", "Username", "Email",
            "Attempt", "Score", "Raw Score", "Status",
            "Start Time", "End Time"
        ]
        ws.append(headers)

        for ue in user_exams:
            ws.append([
                ue.user.id,
                ue.user.username,
                ue.user.email,
                ue.attempt_number,
                ue.score,
                ue.raw_score,
                ue.status,
                ue.start_time.strftime("%Y-%m-%d %H:%M") if ue.start_time else "",
                ue.end_time.strftime("%Y-%m-%d %H:%M") if ue.end_time else "",
            ])

        for col in ws.columns:
            l = max(len(str(v.value)) for v in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = l + 2

        resp = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        resp["Content-Disposition"] = f"attachment; filename=exam_{exam.id}_results.xlsx"
        wb.save(resp)

        return resp
    
    @action(
    detail=True,
    methods=["get"],
    url_path="competency-summary",
    permission_classes=[IsExamInstructorOrAssessor, IsAdmin]
    )
    def competency_summary(self, request, pk=None):
        exam = self.get_object()

        # Ambil semua attempt yang selesai
        user_exams = UserExam.objects.filter(exam=exam, status="completed")

        # Inisialisasi hasil per kategori
        result = {}

        for ue in user_exams:
            answers = UserAnswer.objects.filter(user_exam=ue).select_related("question")

            for ans in answers:
                category = ans.question.category or "uncategorized"
                score = ans.score or 0

                if category not in result:
                    result[category] = {
                        "total_score": 0,
                        "count": 0,
                    }

                result[category]["total_score"] += score
                result[category]["count"] += 1

        summary = []
        for category, data in result.items():
            avg = data["total_score"] / data["count"]

            if avg < 1.5:
                level = "Belum"
            elif avg < 2.5:
                level = "Cukup"
            elif avg < 3.5:
                level = "Menguasai"
            else:
                level = "Sangat Menguasai"

            summary.append({
                "category": category,
                "average_score": round(avg, 2),
                "level": level,
            })

        return Response(summary)



# ================================================================
# TASK VIEWSET
# ================================================================
class CourseTaskViewSet(viewsets.ModelViewSet):
    queryset = CourseTask.objects.all().order_by("-created_at")
    serializer_class = CourseTaskSerializer
    permission_classes = [drf_permissions.IsAuthenticated]

    def get_permissions(self):
        if self.action in ["create", "update", "partial_update", "destroy"]:
            return [IsTrainerOrAdmin()]
        if self.action == "submit_task":
            return [IsCourseParticipant()]
        return [drf_permissions.IsAuthenticated()]


    @action(detail=True, methods=["post"], url_path="submit")
    def submit_task(self, request, pk=None):
        task = self.get_object()

        # must be participant
        if not CourseParticipant.objects.filter(course=task.course, user=request.user).exists():
            return Response({"detail": "Tidak diizinkan."}, status=403)

        # check existing submission
        existing = CourseTaskSubmission.objects.filter(task=task, user=request.user).first()

        if existing:
            # REPLACE submission
            CourseTaskSubmissionFile.objects.filter(submission=existing).delete()

            existing.remarks = request.data.get("remarks", "")
            existing.submitted_at = timezone.now()
            existing.save()

            for f in request.FILES.getlist("files"):
                CourseTaskSubmissionFile.objects.create(submission=existing, file=f)

            return Response({
                "detail": "Submission diperbarui.",
                "submission_id": existing.id,
                "submission": CourseTaskSubmissionSerializer(existing).data
            })

        # FIRST submission
        sub = CourseTaskSubmission.objects.create(
            task=task,
            user=request.user,
            remarks=request.data.get("remarks", "")
        )

        for f in request.FILES.getlist("files"):
            CourseTaskSubmissionFile.objects.create(submission=sub, file=f)

        return Response({
            "detail": "Submit berhasil.",
            "submission_id": sub.id,
            "submission": CourseTaskSubmissionSerializer(sub).data
        }, status=201)


    
    @action(detail=True, methods=["get"], url_path="my")
    def my_submission(self, request, pk=None):
        task = self.get_object()

        sub = CourseTaskSubmission.objects.filter(task=task, user=request.user).first()

        if not sub:
            # FIX: return valid JSON, not empty response
            return Response({}, status=200)

        ser = CourseTaskSubmissionSerializer(sub)
        return Response(ser.data)





# ================================================================
# TASK SUBMISSION VIEWSET
# ================================================================
class TaskSubmissionViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = CourseTaskSubmission.objects.all().order_by("-submitted_at")
    serializer_class = CourseTaskSubmissionSerializer
    permission_classes = [drf_permissions.IsAuthenticated]

    def get_permissions(self):
        # gunakan self.request, bukan request local yang tidak didefinisikan
        if self.action in ["list", "retrieve"]:
            # admin lihat semua, trainer/assessor juga boleh
            if self.request.user.is_staff:
                return [IsAdmin()]
            return [IsTrainerOrAssessor()]
        return [drf_permissions.IsAuthenticated()]


class AdminDashboardAPIView(APIView):
    """
    GET /api/exam/dashboard/admin/
    Return aggregated data used by Admin Dashboard.
    """
    permission_classes = [drf_permissions.IsAuthenticated, IsAdmin]

    def get(self, request):
        today = timezone.localdate()

        # Basic counts
        total_courses = Course.objects.count()
        total_participants = CourseParticipant.objects.filter(role="participant").count()
        total_trainers = CourseParticipant.objects.filter(role="trainer").values("user").distinct().count()
        total_assessors = CourseParticipant.objects.filter(role="assessor").values("user").distinct().count()
        total_exams = Exam.objects.count()
        total_tasks = CourseTask.objects.count()

        # Pending requirement approvals
        pending_requirements = CourseRequirementSubmission.objects.filter(status="pending").count()

        # Pending task grading
        pending_task_grading = CourseTaskSubmission.objects.filter(graded=False).count()


        # Pending essay grading
        pending_essay_grading = UserAnswer.objects.filter(text_answer__isnull=False).exclude(text_answer="").filter(graded=False).count()

        # Running / active courses (overlapping today)
        running_courses_qs = Course.objects.filter(start_date__lte=today, end_date__gte=today).order_by("start_date")[:10]
        running_courses = [
            {"id": c.id, "title": c.title, "start_date": c.start_date, "end_date": c.end_date}
            for c in running_courses_qs
        ]

        # Active exams (now between start and end time) - show those with end_time >= now
        now = timezone.now()
        active_exams_qs = Exam.objects.filter(start_time__lte=now, end_time__gte=now).select_related("course")[:10]
        active_exams = [
            {
                "id": e.id,
                "title": e.title,
                "course_title": e.course.title if e.course else None,
                "end_time": e.end_time
            }
            for e in active_exams_qs
        ]

        # Tasks due today
        tasks_due_today_qs = CourseTask.objects.filter(due_date=today).select_related("course")[:10]
        tasks_due_today = [
            {"id": t.id, "course_title": t.course.title if t.course else None, "title": t.title, "due_date": t.due_date}
            for t in tasks_due_today_qs
        ]

        data = {
            "stats": {
                "total_courses": total_courses,
                "total_participants": total_participants,
                "total_trainers": total_trainers,
                "total_assessors": total_assessors,
                "total_exams": total_exams,
                "total_tasks": total_tasks,
                "pending_requirements": pending_requirements,
                "pending_task_grading": pending_task_grading,
                "pending_essay_grading": pending_essay_grading
            },
            "running_courses": running_courses,
            "active_exams": active_exams,
            "today_deadlines": tasks_due_today
        }

        return Response(data, status=status.HTTP_200_OK)